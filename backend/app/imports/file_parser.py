"""Reads an uploaded CSV/XLS/XLSX file into a raw grid (list of rows, cells
kept in their native type) — deliberately no pandas/typed parsing here, so
header detection (app/imports/detection.py) sees the file exactly as it
is before any type inference could obscure where the real header row is.

Only ever reads a bounded window of rows: detection needs at most ~15 rows
to find the header plus ~200 more to sample columns, regardless of how
large the underlying file is.
"""

import csv
import io
import itertools
from datetime import date, datetime

import openpyxl
import xlrd

from app.imports.exceptions import UnsupportedFileType

Row = list[object]


def read_rows(file_bytes: bytes, filename: str, *, start: int = 0, max_rows: int = 200) -> list[Row]:
    extension = filename.lower().rsplit(".", 1)[-1] if "." in filename else ""
    if extension == "csv":
        rows = _read_csv_rows(file_bytes, start=start, max_rows=max_rows)
    elif extension == "xlsx":
        rows = _read_xlsx_rows(file_bytes, start=start, max_rows=max_rows)
    elif extension == "xls":
        rows = _read_xls_rows(file_bytes, start=start, max_rows=max_rows)
    else:
        raise UnsupportedFileType(extension)

    width = max((len(row) for row in rows), default=0)
    return [row + [None] * (width - len(row)) for row in rows]


def _decode_csv_bytes(file_bytes: bytes) -> str:
    for encoding in ("utf-8-sig", "cp1252"):
        try:
            return file_bytes.decode(encoding)
        except UnicodeDecodeError:
            continue
    return file_bytes.decode("utf-8", errors="replace")


def _read_csv_rows(file_bytes: bytes, *, start: int, max_rows: int) -> list[Row]:
    text = _decode_csv_bytes(file_bytes)
    reader = csv.reader(io.StringIO(text))
    return [list(row) for row in itertools.islice(reader, start, start + max_rows)]


def _read_xlsx_rows(file_bytes: bytes, *, start: int, max_rows: int) -> list[Row]:
    workbook = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    try:
        sheet = workbook.active
        row_iter = sheet.iter_rows(values_only=True)
        return [list(row) for row in itertools.islice(row_iter, start, start + max_rows)]
    finally:
        workbook.close()


def _read_xls_rows(file_bytes: bytes, *, start: int, max_rows: int) -> list[Row]:
    # xlrd loads the whole legacy-format sheet regardless of any window we
    # ask for — a real limitation of the .xls format, not this code.
    workbook = xlrd.open_workbook(file_contents=file_bytes)
    sheet = workbook.sheet_by_index(0)
    end = min(sheet.nrows, start + max_rows)
    rows = []
    for i in range(start, end):
        row = sheet.row_values(i)
        rows.append([_xls_cell(sheet, i, j) for j in range(len(row))])
    return rows


def _xls_cell(sheet, row_index: int, col_index: int) -> object:
    # xlrd represents dates as raw floats + a separate cell type — convert
    # to a real datetime so downstream code sees the same native types it
    # would from openpyxl/csv, per file_parser's contract.
    cell = sheet.cell(row_index, col_index)
    if cell.ctype == xlrd.XL_CELL_DATE:
        tup = xlrd.xldate_as_tuple(cell.value, sheet.book.datemode)
        return datetime(*tup) if any(tup[3:]) else date(*tup[:3])
    return cell.value


def normalize_cell(value: object) -> str:
    """Display/comparison form for header-row scoring — never used for the
    actual mapped data (that stays in its native type for value_parsers.py).
    """
    if value is None:
        return ""
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()
